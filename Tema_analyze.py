from flask import Flask, render_template, request
import requests
import pandas as pd
import os
import getpass
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/result', methods=['POST'])
def result():
    global Team
    Team = request.form['team']
   
    if sys.platform == 'win32':
        username = getpass.getuser()
        Pdata_file_url = 'https://drive.google.com/uc?id=13vVpc0a87eyqUDyISthnx4TP8XfrAmyB'
        Tdata_file_url = 'https://drive.google.com/uc?id=1gEshT9oOmnEbrqfuSYn9T09BFfokH1Ye'

        output_path1 = 'C:/Users/' + username + '/Documents/Pdata_file.csv'
        output_path2 = 'C:/Users/' + username + '/Documents/Tdata_file.csv'

    elif sys.platform == 'darwin':
        username = getpass.getuser()
        Pdata_file_url = 'https://drive.google.com/uc?id=13vVpc0a87eyqUDyISthnx4TP8XfrAmyB'
        Tdata_file_url = 'https://drive.google.com/uc?id=1gEshT9oOmnEbrqfuSYn9T09BFfokH1Ye'

        output_path1 = '/Users/' + username + '/Documents/Pdata_file.csv'
        output_path2 = '/Users/' + username + '/Documents/Tdata_file.csv'

    def download_file(url, output_path):
        response = requests.get(url)
        with open(output_path, 'wb') as file:
            file.write(response.content)

    download_file(Pdata_file_url, output_path1)
    download_file(Tdata_file_url, output_path2)

    if os.path.exists(Pdata_file_url):
        new_wb = load_workbook(Pdata_file_url)
    else:
        new_wb = Workbook()

    if os.path.exists(Tdata_file_url):
        new_wb = load_workbook(Tdata_file_url)
    else:
        new_wb = Workbook()

    P_data = pd.read_csv(output_path1, sep =',')
    T_data = pd.read_csv(output_path2, sep =',')

    P_namedata = P_data[['NAME', 'TEAM']].copy() #이름과 팀이 저장된 데이터 프레임 생성
    P_stats = []
    max_stats = []

    result_data = P_namedata.iloc[:10, :3]
    result_Teamdata = T_data.loc[T_data['TEAM_NAME'] == Team]

    return render_template('result.html', team=result_Teamdata)

global Team

@app.route('/result2', methods=['POST'])
def result2():
    stats = request.form.getlist('stats')
    global Team

    if sys.platform == 'win32':
        username = getpass.getuser()
        Pdata_file_url = 'https://drive.google.com/uc?id=13vVpc0a87eyqUDyISthnx4TP8XfrAmyB'
        Tdata_file_url = 'https://drive.google.com/uc?id=1gEshT9oOmnEbrqfuSYn9T09BFfokH1Ye'

        output_path1 = 'C:/Users/' + username + '/Documents/Pdata_file.csv'
        output_path2 = 'C:/Users/' + username + '/Documents/Tdata_file.csv'

    elif sys.platform == 'darwin':
        username = getpass.getuser()
        Pdata_file_url = 'https://drive.google.com/uc?id=13vVpc0a87eyqUDyISthnx4TP8XfrAmyB'
        Tdata_file_url = 'https://drive.google.com/uc?id=1gEshT9oOmnEbrqfuSYn9T09BFfokH1Ye'

        output_path1 = '/Users/' + username + '/Documents/Pdata_file.csv'
        output_path2 = '/Users/' + username + '/Documents/Tdata_file.csv'

    def download_file(url, output_path):
        response = requests.get(url)
        with open(output_path, 'wb') as file:
            file.write(response.content)

    download_file(Pdata_file_url, output_path1)
    download_file(Tdata_file_url, output_path2)

    P_data = pd.read_csv(output_path1, sep =',')
    T_data = pd.read_csv(output_path2, sep =',')

    P_namedata = P_data[['NAME', 'TEAM']].copy() #이름과 팀이 저장된 데이터 프레임 생성
    P_stats = []
    max_stats = []

    def arrange(stats): #배열 및 점수 부여 함수
        selected_cols = ['NAME', stats]
        selected_data = P_data[selected_cols]
        if stats == 'TOV' or stats == 'PF':
            sorted_data = selected_data.sort_values(by=stats, ascending=True)
        else:
            sorted_data = selected_data.sort_values(by=stats, ascending=False)
        sorted_data.reset_index(drop=True, inplace=True) #정렬 및 인덱스 초기화
        sorted_data[stats] = sorted_data[stats].rank(method='max', ascending=True) #점수 부여, 동점자는 동일한 점수
        return sorted_data[['NAME', stats]]

    P_stats = stats

    for i in range(len(P_stats)): #점수가 포함된 데이터 프레임 생성, 점수의 개수만큼 반복
        merged_data = pd.merge(P_namedata, arrange(P_stats[i]), on='NAME', how='outer')
        P_namedata = merged_data

    P_namedata['sum_score'] = P_namedata.iloc[:, 2:].sum(axis=1) #모든 점수 값 합산
    P_namedata = P_namedata.sort_values(by='sum_score', ascending=False) #합산 값 토대로 내림차순 정렬

    max_stats = P_namedata.iloc[:, 2:-1].idxmax(axis=1) #각 점수별 최대값 탐색
    melted_data = P_namedata.melt(id_vars=['NAME', 'TEAM'], value_vars=max_stats, var_name='stat', value_name='max_value') 
    P_namedata['max'] = melted_data.loc[melted_data.groupby(['NAME'])['max_value'].idxmax(), 'stat']
    P_namedata.insert(2, 'max_stats', max_stats)

    condition = (P_namedata['TEAM'] != Team) #팀 값 입력받고, 그 팀에 존재하는 선수일경우 데이터 프레임에서 제외
    P_namedata = P_namedata[condition]

    P_namedata.reset_index(drop=True, inplace=True)
    P_namedata.index = P_namedata.index + 1

    result_data = P_namedata.iloc[:10, :3]
    result_Teamdata = T_data.loc[T_data['TEAM_NAME'] == Team]

    return render_template('result2.html', team=result_Teamdata, result_data=result_data)

if __name__ == '__main__':
    app.run(debug=True)